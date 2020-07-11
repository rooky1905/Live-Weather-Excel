import requests
import openpyxl
import time

wb = openpyxl.load_workbook('test1.xlsx')
sheet_obj = wb.active 

while True:
    rows = sheet_obj.max_row

    for i in range(2, rows+1): 
        cell_obj = sheet_obj.cell(row = i, column = 1) 
        temp = sheet_obj.cell(row = i, column = 2)
        update = sheet_obj.cell(row = i, column = 3)
        if update.value == 1:
            if temp.value == 'F':
                unit = 'imperial'
            else:
                unit = 'metric'
            response = requests.get('http://api.openweathermap.org/data/2.5/weather?apikey=4a2360d14bf33378079d2e2d49e35ddb&mode=json&units={}&q={}'.format(unit,cell_obj.value))
            sheet_obj.cell(column=4 , row=i, value=response.json()['main']['temp'])
        else:
            continue
    wb.save('test1.xlsx')
    print("Info Updated...")
    time.sleep(3)

#File gets updated every 3 second.